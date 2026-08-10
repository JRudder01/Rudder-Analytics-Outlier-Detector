import io
import re
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from sklearn.linear_model import HuberRegressor


DEFAULT_HEADERS = ["Close Price", "Living Area"]


def clean_number(value):
    """Convert common Excel-style numeric text into a float."""
    text = str(value).strip()
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    text = re.sub(r"[$,%\s]", "", text).replace(",", "")

    try:
        number = float(text)
        return -number if negative else number
    except ValueError:
        return np.nan


def _split_pasted_line(line: str) -> List[str]:
    """Split one row copied from Excel into exactly two cells."""
    if "\t" in line:
        return [part.strip() for part in line.split("\t") if part.strip()]

    # Fallback if tabs were converted to multiple spaces.
    parts = [part.strip() for part in re.split(r"\s{2,}", line.strip()) if part.strip()]
    if len(parts) == 2:
        return parts

    match = re.match(r"^(.*?)\s+([^\s]+)$", line.strip())
    if match:
        return [match.group(1).strip(), match.group(2).strip()]

    return parts


def parse_pasted_data(
    raw_text: str,
    first_row_mode: str = "auto",
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Parse two columns copied from Excel.

    first_row_mode: "auto", "header", or "data".
    """
    lines = [line for line in raw_text.strip().splitlines() if line.strip()]
    if not lines:
        raise ValueError("Paste the Close Price and Living Area columns first.")

    rows = []
    for line_number, line in enumerate(lines, start=1):
        row = _split_pasted_line(line)
        if len(row) != 2:
            raise ValueError(
                f"Pasted row {line_number} does not contain exactly two columns. "
                "Copy both Excel columns together and paste them into the box."
            )
        rows.append(row)

    first_row_is_numeric = all(pd.notna(clean_number(value)) for value in rows[0])

    if first_row_mode == "auto":
        has_header = not first_row_is_numeric
    elif first_row_mode == "header":
        has_header = True
    elif first_row_mode == "data":
        has_header = False
    else:
        raise ValueError("Invalid first-row setting.")

    if has_header:
        headers = rows[0]
        rows = rows[1:]
    else:
        headers = DEFAULT_HEADERS.copy()

    if not rows:
        raise ValueError("No listing rows were found beneath the headers.")

    df = pd.DataFrame(rows, columns=DEFAULT_HEADERS)
    df["Close Price"] = df["Close Price"].map(clean_number)
    df["Living Area"] = df["Living Area"].map(clean_number)

    invalid = df[
        df["Close Price"].isna()
        | df["Living Area"].isna()
        | (df["Close Price"] <= 0)
        | (df["Living Area"] <= 0)
    ]

    if not invalid.empty:
        pasted_rows = ", ".join(str(i + 1 + (1 if has_header else 0)) for i in invalid.index[:10])
        raise ValueError(
            "One or more rows contain missing, text, zero, or negative values. "
            f"Check pasted row(s): {pasted_rows}."
        )

    return df, headers


def robust_z(values: np.ndarray) -> np.ndarray:
    values = np.asarray(values, dtype=float)
    median = np.median(values)
    mad = np.median(np.abs(values - median))

    if mad > 1e-12:
        return 0.67448975 * (values - median) / mad

    std = np.std(values, ddof=1) if len(values) > 1 else 0
    if std > 1e-12:
        return (values - np.mean(values)) / std

    return np.zeros_like(values)


def analyze_outliers(
    df: pd.DataFrame,
    robust_z_threshold: float = 3.5,
    iqr_multiplier: float = 1.5,
    outlier_method: str = "either",
    flag_extreme_living_area: bool = False,
) -> Tuple[pd.DataFrame, Dict[str, float]]:
    """Run the same outlier logic used in the Colab notebook."""
    if len(df) < 3:
        raise ValueError("At least 3 listings are required for the analysis.")

    results = df.copy()
    results["Price per Sq Ft"] = results["Close Price"] / results["Living Area"]

    x = np.log(results["Living Area"].to_numpy(float)).reshape(-1, 1)
    y = np.log(results["Close Price"].to_numpy(float))

    model = HuberRegressor(epsilon=1.35, max_iter=1000)
    try:
        model.fit(x, y)
        predicted_log = model.predict(x)
    except Exception:
        slope, intercept = np.polyfit(x.ravel(), y, 1)
        predicted_log = intercept + slope * x.ravel()

    results["Expected Close Price"] = np.exp(predicted_log)
    results["Difference from Expected"] = (
        results["Close Price"] / results["Expected Close Price"] - 1
    )
    results["Regression Robust Z"] = robust_z(y - predicted_log)
    results["Regression Outlier"] = (
        results["Regression Robust Z"].abs() > robust_z_threshold
    )

    ppsf = results["Price per Sq Ft"]
    ppsf_q1, ppsf_q3 = ppsf.quantile([0.25, 0.75])
    ppsf_iqr = ppsf_q3 - ppsf_q1
    ppsf_low = ppsf_q1 - iqr_multiplier * ppsf_iqr
    ppsf_high = ppsf_q3 + iqr_multiplier * ppsf_iqr
    results["Price/Sq Ft Outlier"] = (ppsf < ppsf_low) | (ppsf > ppsf_high)

    area = results["Living Area"]
    area_q1, area_q3 = area.quantile([0.25, 0.75])
    area_iqr = area_q3 - area_q1
    area_low = area_q1 - iqr_multiplier * area_iqr
    area_high = area_q3 + iqr_multiplier * area_iqr
    results["Extreme Living Area"] = (area < area_low) | (area > area_high)

    if outlier_method == "robust_regression":
        results["Outlier"] = results["Regression Outlier"]
    elif outlier_method == "price_per_sqft_iqr":
        results["Outlier"] = results["Price/Sq Ft Outlier"]
    elif outlier_method == "either":
        results["Outlier"] = (
            results["Regression Outlier"] | results["Price/Sq Ft Outlier"]
        )
    else:
        raise ValueError("Invalid outlier method.")

    if flag_extreme_living_area:
        results["Outlier"] = results["Outlier"] | results["Extreme Living Area"]

    diagnostics = {
        "listings": int(len(results)),
        "outliers": int(results["Outlier"].sum()),
        "regression_outliers": int(results["Regression Outlier"].sum()),
        "ppsf_outliers": int(results["Price/Sq Ft Outlier"].sum()),
        "ppsf_low": float(ppsf_low),
        "ppsf_high": float(ppsf_high),
    }

    return results, diagnostics


def build_excel(results: pd.DataFrame, headers: List[str]) -> bytes:
    """
    Build a one-sheet workbook containing every listing in original order.
    Outliers receive only a blue cell fill.
    """
    output = results[["Close Price", "Living Area"]].round(0).astype(int).copy()
    output.columns = headers

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        output.to_excel(writer, sheet_name="Highlighted_Output", index=False)
        ws = writer.sheets["Highlighted_Output"]

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(
            name="Aptos Narrow",
            size=11,
            color="FFFFFF",
            bold=True,
        )
        data_font = Font(
            name="Aptos Narrow",
            size=11,
            bold=False,
        )
        outlier_fill = PatternFill(fill_type="solid", fgColor="00B0F0")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for excel_row in range(2, ws.max_row + 1):
            price_cell = ws.cell(excel_row, 1)
            area_cell = ws.cell(excel_row, 2)

            price_cell.number_format = "#,##0"
            area_cell.number_format = "0"
            price_cell.font = data_font
            area_cell.font = data_font

        for dataframe_index in results.index[results["Outlier"]]:
            excel_row = int(dataframe_index) + 2
            ws.cell(excel_row, 1).fill = outlier_fill
            ws.cell(excel_row, 2).fill = outlier_fill

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 16

    buffer.seek(0)
    return buffer.getvalue()
